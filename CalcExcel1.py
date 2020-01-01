import pathlib
import openpyxl
import csv


lwb = openpyxl.Workbook()
lsh = lwb.active

list_raw = 1
sum_score = 0
path = pathlib.Path("..\Python_PRG")
for pass_obj in path.iterdir():
    if pass_obj.match("*.xlsx"):
        wb = openpyxl.load_workbook(pass_obj)
        for sh in wb:
            for dt_row in range(2,12):
                if sh.cell(dt_row, 2).value != None:
                    sum_score += sh.cell(dt_row, 2).value        
print (sum_score)
